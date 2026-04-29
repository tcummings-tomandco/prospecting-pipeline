import argparse, os, sys, time, json, requests, pandas as pd
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = "https://api.apollo.io/api/v1"

# Role profiles — each defines who to target.
# Primary = roles to try first. Fallback = broader decision-makers used if primary returns nothing.
ROLE_PROFILES = {
    "Ecommerce Decision Makers": {
        "description": "Ecommerce Managers, Directors, Heads of Digital — for retail/DTC prospecting.",
        "primary": [
            "ecommerce manager", "ecommerce director", "head of ecommerce",
            "vp ecommerce", "director of ecommerce", "digital director",
            "head of digital", "director of digital", "chief digital officer",
            "chief marketing officer", "head of online", "online director",
            "vp digital", "director of online sales",
        ],
        "fallback": [
            "managing director", "owner", "founder", "ceo", "general manager",
            "marketing manager", "head of marketing", "marketing director",
            "commercial director", "operations director", "sales director",
        ],
    },
    "AI Agency Targeting": {
        "description": "Managing Partners, Heads of Technical, Ops Directors — for B2B agency/consulting prospects.",
        "primary": [
            "managing partner", "head of technical", "head of technology",
            "operations director", "head of operations", "chief operating officer",
            "chief technology officer", "head of ai", "head of innovation",
            "innovation director", "transformation director", "head of digital transformation",
            "vp engineering", "head of engineering", "director of operations",
        ],
        "fallback": [
            "managing director", "owner", "founder", "ceo", "partner",
            "general manager", "director", "principal",
        ],
    },
    "Marketing & Brand": {
        "description": "CMOs, Marketing Directors, Heads of Brand — for marketing-led pitches.",
        "primary": [
            "chief marketing officer", "marketing director", "head of marketing",
            "vp marketing", "director of marketing", "head of brand",
            "brand director", "head of growth", "growth director",
            "head of communications", "communications director",
        ],
        "fallback": [
            "managing director", "owner", "founder", "ceo", "general manager",
            "head of digital", "commercial director",
        ],
    },
    "Senior Decision Makers (General)": {
        "description": "Founders, CEOs, MDs, Directors — broad C-suite/leadership across any industry.",
        "primary": [
            "ceo", "managing director", "founder", "co-founder", "owner",
            "president", "managing partner", "chief executive officer",
        ],
        "fallback": [
            "director", "vp", "chief operating officer", "general manager",
            "head of", "partner",
        ],
    },
}

DEFAULT_PROFILE = "Ecommerce Decision Makers"

SENIORITIES = ["owner", "founder", "c_suite", "vp", "director", "manager"]


def _headers():
    key = os.environ.get("APOLLO_API_KEY", "")
    if not key:
        raise RuntimeError("APOLLO_API_KEY not set. Add it to your .env file or environment.")
    return {"x-api-key": key, "Content-Type": "application/json"}


def clean_domain(website):
    """Extract first clean domain from website field."""
    if pd.isna(website):
        return None
    w = str(website).strip()
    for prefix in ["https://www.", "http://www.", "https://", "http://", "www."]:
        w = w.replace(prefix, "")
    w = w.split("/")[0].strip()
    for sep in [" / ", " /", "/ ", " "]:
        if sep in w:
            w = w.split(sep)[0].strip()
            break
    if "(" in w:
        w = w.split("(")[0].strip()
    return w if w and "." in w else None


def search_people(domain, primary_titles, fallback_titles):
    """Search Apollo for matching titles at a domain, with fallback."""
    payload = {
        "q_organization_domains_list": [domain],
        "person_titles": primary_titles,
        "person_seniorities": SENIORITIES,
        "page": 1,
        "per_page": 10,
    }
    try:
        r = requests.post(f"{BASE}/mixed_people/api_search", headers=_headers(), json=payload)
        r.raise_for_status()
        people = r.json().get("people", [])
        if people:
            return people, "primary"

        if not fallback_titles:
            return [], "fallback"

        time.sleep(0.3)
        payload["person_titles"] = fallback_titles
        r = requests.post(f"{BASE}/mixed_people/api_search", headers=_headers(), json=payload)
        r.raise_for_status()
        return r.json().get("people", []), "fallback"
    except requests.exceptions.RequestException as e:
        print(f"[APOLLO SEARCH] error for {domain}: {e}", flush=True)
        return [], f"error: {e}"


def enrich_person(person):
    """Enrich a person to get their email and full name."""
    payload = {"id": person["id"], "reveal_personal_emails": True}
    try:
        r = requests.post(f"{BASE}/people/match", headers=_headers(), json=payload)
        r.raise_for_status()
        p = r.json().get("person", {})
        return {
            "first_name": p.get("first_name", person.get("first_name", "")),
            "last_name": p.get("last_name", person.get("last_name", "")),
            "email": p.get("email", ""),
            "email_status": p.get("email_status", ""),
            "phone": (p.get("phone_numbers") or [{}])[0].get("sanitized_number", "") if p.get("phone_numbers") else "",
            "linkedin": p.get("linkedin_url", ""),
        }
    except requests.exceptions.RequestException as e:
        print(f"[APOLLO ENRICH] error for {person.get('id', 'unknown')}: {e}", flush=True)
        return {"first_name": person.get("first_name", ""), "last_name": person.get("last_name", ""),
                "email": "", "email_status": "", "phone": "", "linkedin": ""}


def _find_column(df, candidates):
    """Find the first matching column name (case-insensitive)."""
    col_map = {c.lower().strip(): c for c in df.columns}
    for name in candidates:
        if name.lower() in col_map:
            return col_map[name.lower()]
    return None


def enrich_companies(df, category, on_progress=None, contacts_per_company=3,
                     primary_titles=None, fallback_titles=None, profile_name=None):
    """Enrich a DataFrame of companies. Calls on_progress(idx, total, company, message) per step.

    Targeting:
    - profile_name: name of a preset in ROLE_PROFILES (used if primary_titles not given)
    - primary_titles / fallback_titles: explicit lists of titles to target (override profile)

    Returns a DataFrame of results.
    """
    # Resolve targeting
    if primary_titles is None:
        profile = ROLE_PROFILES.get(profile_name or DEFAULT_PROFILE, ROLE_PROFILES[DEFAULT_PROFILE])
        primary_titles = profile["primary"]
        if fallback_titles is None:
            fallback_titles = profile["fallback"]

    if fallback_titles is None:
        fallback_titles = []

    # Flexible column matching
    company_col = _find_column(df, ["Company", "Company Name", "Name", "Organisation", "Organization"])
    website_col = _find_column(df, ["Website", "Domain", "URL", "Web", "Site"])

    if not company_col:
        raise ValueError(f"No 'Company' column found. Columns in file: {list(df.columns)}")
    if not website_col:
        raise ValueError(f"No 'Website' column found. Columns in file: {list(df.columns)}")

    results = []
    total = len(df)

    def log(idx, company, message):
        if on_progress:
            on_progress(idx, total, company, message)

    for idx, row in df.iterrows():
        company = row[company_col]
        domain = clean_domain(row[website_col])
        log(idx, company, f"Processing {company} ({domain or 'no domain'})")

        if not domain:
            results.append({
                "Company": company, "Website": row[website_col],
                "First Name": "", "Last Name": "", "Title": "", "Email": "",
                "Email Status": "", "Phone": "", "LinkedIn": "",
                "Search Type": "no domain", "Category": category,
            })
            continue

        people, search_type = search_people(domain, primary_titles, fallback_titles)
        time.sleep(0.3)

        if not people:
            log(idx, company, f"  No contacts found for {company}")
            results.append({
                "Company": company, "Website": row[website_col],
                "First Name": "", "Last Name": "", "Title": "", "Email": "",
                "Email Status": "", "Phone": "", "LinkedIn": "",
                "Search Type": "none found", "Category": category,
            })
            continue

        label = "ecommerce" if search_type == "primary" else "fallback"
        log(idx, company, f"  Found {len(people)} via {label}, enriching top {contacts_per_company}...")

        for person in people[:contacts_per_company]:
            title = person.get("title", "")
            enriched = enrich_person(person)
            time.sleep(0.3)

            first = enriched["first_name"]
            last = enriched["last_name"]
            results.append({
                "Company": company,
                "Website": row[website_col],
                "First Name": first,
                "Last Name": last,
                "Title": title,
                "Email": enriched["email"],
                "Email Status": enriched["email_status"],
                "Phone": enriched["phone"],
                "LinkedIn": enriched["linkedin"],
                "Search Type": search_type,
                "Category": category,
            })
            log(idx, company, f"    {first} {last} - {title} - {enriched['email']}")

    return pd.DataFrame(results)


def save_enriched_excel(df, output_file):
    """Save enriched DataFrame to a formatted Excel file."""
    df.to_excel(output_file, index=False, sheet_name="Enriched Contacts")

    wb = load_workbook(output_file)
    ws = wb.active

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    cell_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = cell_font
            cell.border = thin_border

    widths = {"A": 28, "B": 35, "C": 18, "D": 18, "E": 35, "F": 35, "G": 14, "H": 18, "I": 40, "J": 16, "K": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(output_file)


def main():
    parser = argparse.ArgumentParser(description="Enrich company list with Apollo contacts")
    parser.add_argument("input_file", help="Excel file with Company and Website columns")
    parser.add_argument("--category", "-c", default=None,
                        help="Category label (e.g. 'Furniture'). Defaults to filename stem.")
    parser.add_argument("--profile", "-p", default=DEFAULT_PROFILE,
                        choices=list(ROLE_PROFILES.keys()),
                        help="Role profile to target. Default: 'Ecommerce Decision Makers'.")
    args = parser.parse_args()

    input_file = args.input_file
    category = args.category or os.path.splitext(os.path.basename(input_file))[0].replace(" Companies", "")
    output_file = input_file.replace(".xlsx", " - Enriched.xlsx")

    df = pd.read_excel(input_file)
    print(f"Loaded {len(df)} companies from {input_file}")
    print(f"Category: {category}")
    print(f"Profile: {args.profile}\n")

    def on_progress(idx, total, company, message):
        print(f"[{idx+1}/{total}] {message}")

    result_df = enrich_companies(df, category, on_progress=on_progress, profile_name=args.profile)
    save_enriched_excel(result_df, output_file)

    print(f"\nDone! {len(result_df)} rows written to {output_file}")
    verified = (result_df["Email Status"] == "verified").sum()
    print(f"Verified emails: {verified}")


if __name__ == "__main__":
    main()
